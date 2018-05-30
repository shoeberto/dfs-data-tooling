from openpyxl import Workbook
import dfs.datasheets.datasheet as datasheet


class DatasheetWriter:
    def write(self, sheet, output_directory):
        """
        Write a datasheet to an xlsx file.

        Keyword arguments:
        sheet -- a datasheet object.
        output_directory -- the directory to write the new xlsx file to.
        """

        # TODO: make this support treatment datasheets as well as plot datasheets
        workbook = Workbook()

        # remove the default worksheet
        workbook.remove(workbook['Sheet'])

        general_tab = workbook.create_sheet(title=datasheet.TAB_NAME_GENERAL)
        self.format_general_tab(sheet, general_tab)

        witness_tree_tab = workbook.create_sheet(title=datasheet.TAB_NAME_WITNESS_TREES)
        self.format_witness_trees_tab(sheet, witness_tree_tab)

        notes_tab = workbook.create_sheet(title=datasheet.TAB_NAME_NOTES)
        self.format_notes_tab(sheet, notes_tab)

        tree_table_tab = workbook.create_sheet(title=datasheet.TAB_NAME_TREE_TABLE)
        self.format_tree_table_tab(sheet, tree_table_tab)

        cover_table_tab = workbook.create_sheet(title=datasheet.TAB_NAME_COVER_TABLE)
        self.format_cover_table_tab(sheet, cover_table_tab)

        sapling_tab = workbook.create_sheet(title=datasheet.TAB_NAME_SAPLING)
        self.format_sapling_tab(sheet, sapling_tab)

        seedling_tab = workbook.create_sheet(title=datasheet.TAB_NAME_SEEDLING)
        self.format_seedling_tab(sheet, seedling_tab)

        workbook.save('{}/{}'.format(output_directory, sheet.input_filename))


    def format_general_tab(self, sheet, tab):
        tab['A1'] = 'Study Area'
        tab['B1'] = str(sheet.tabs[datasheet.TAB_NAME_GENERAL].study_area)

        tab['A2'] = 'Plot Number'
        tab['B2'] = str(sheet.tabs[datasheet.TAB_NAME_GENERAL].plot_number)

        tab['A3'] = 'Deer Impact'
        tab['B3'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].deer_impact

        tab['A4'] = 'Date'
        tab['B4'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].collection_date
        tab['B4'].number_format = 'M/D/YYYY'

        tab['A6'] = 'Coordinate Converter'

        tab['A7'] = 'INPUT LAT HERE'
        tab['B7'] = 'INPUT LONG HERE'
        tab['C7'] = 'Data Type'
        tab['D7'] = 'Yes/No'
        tab['E7'] = 'Yes/No'
        tab['F7'] = '0-359'
        tab['G7'] = 'Feet'
        tab['H7'] = 'Meters'
        tab['I7'] = 'DO NOT INPUT VALUES HERE'

        tab['A8'] = 'ex. 4045.1291'
        tab['B8'] = 'ex. 7743.2763'
        tab['C8'] = 'Sub/ Micro'
        tab['D8'] = 'Collected'
        tab['E8'] = 'Fenced'
        tab['F8'] = 'Azimuth'
        tab['G8'] = 'Distance'
        tab['H8'] = 'Altitude'
        tab['I8'] = 'Latitude'
        tab['J8'] = 'Longitude'
        tab['K8'] = 'UID'

        i = 0
        for rownumber in range(9, 14):
            subplot = sheet.tabs[datasheet.TAB_NAME_GENERAL].subplots[i]

            tab['A{}'.format(rownumber)] = subplot.latitude
            tab['B{}'.format(rownumber)] = subplot.longitude
            tab['C{}'.format(rownumber)] = str(subplot.micro_plot_id)
            tab['D{}'.format(rownumber)] = subplot.collected
            tab['E{}'.format(rownumber)] = subplot.fenced
            tab['F{}'.format(rownumber)] = subplot.azimuth
            tab['G{}'.format(rownumber)] = subplot.distance
            tab['H{}'.format(rownumber)] = subplot.altitude
            tab['I{}'.format(rownumber)] = '=LEFT((LEFT(A{0},2)+(RIGHT(A{0},LEN(A{0})-2)/60)),10)'.format(rownumber)
            tab['J{}'.format(rownumber)] = '=LEFT(-1*(LEFT(B{0},2)+(RIGHT(B{0},LEN(B{0})-2)/60)),10)'.format(rownumber)
            tab['K{}'.format(rownumber)] = '=VALUE(CONCATENATE(General!$B$1,IF(LEN(General!$B$2)<2, CONCATENATE(0,General!$B$2),General!$B$2),IF(LEN(C{0})<2, CONCATENATE(0,C{0}),C{0}))'.format(rownumber)

            i += 1

        tab['A15'] = 'Data Type'
        tab['B15'] = 'Yes/No'
        tab['C15'] = 'Yes/No'
        tab['D15'] = '0-1'
        tab['E15'] = 'Type'
        tab['F15'] = 'Yes/No'
        tab['G15'] = 'Yes/No'

        tab['A16'] = 'Sub/Micro'
        tab['B16'] = 'Re-Monumented'
        tab['C16'] = 'Forested'
        tab['D16'] = 'Disturbance'
        tab['E16'] = 'Type'
        tab['F16'] = 'Lime'
        tab['G16'] = 'Herbicide'

        i = 0
        for rownumber in range(17, 22):
            subplot = sheet.tabs[datasheet.TAB_NAME_GENERAL].subplots[i]

            tab['A{}'.format(rownumber)] = subplot.micro_plot_id
            tab['B{}'.format(rownumber)] = subplot.re_monumented
            tab['C{}'.format(rownumber)] = subplot.forested
            tab['D{}'.format(rownumber)] = subplot.disturbance
            tab['E{}'.format(rownumber)] = subplot.disturbance_type
            tab['F{}'.format(rownumber)] = subplot.lime
            tab['G{}'.format(rownumber)] = subplot.herbicide

            i += 1

        tab['A23'] = 'Fenced Subplot Condition'
        tab['C23'] = '0-3'

        tab['A24'] = 'Active Exclosure'
        tab['B24'] = 'Repair(s)'
        tab['C24'] = 'Level'
        tab['C24'] = 'Level'

        tab['A25'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].fenced_subplot_condition.active_exclosure
        tab['B25'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].fenced_subplot_condition.repairs
        tab['C25'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].fenced_subplot_condition.level

        tab['A27'] = 'Auxillary Post Locations'
        tab['A28'] = '1-5'
        tab['B28'] = '1 or 2'
        tab['C28'] = 'Wooden or Rebar'
        tab['D28'] = '0-359'
        tab['E28'] = 'Feet'

        tab['A29'] = 'Sub/Micro'
        tab['B29'] = 'Post'
        tab['C29'] = 'Stake Type'
        tab['D29'] = 'Azimuth'
        tab['E29'] = 'Distance'

        i = 0
        for rownumber in range(30, 35):
            if i < len(sheet.tabs[datasheet.TAB_NAME_GENERAL].auxillary_post_locations):
                auxillary_post_location = sheet.tabs[datasheet.TAB_NAME_GENERAL].auxillary_post_location[i]

                tab['A{}'.format(rownumber)] = auxillary_post_location.micro_plot_id
                tab['B{}'.format(rownumber)] = auxillary_post_location.post
                tab['C{}'.format(rownumber)] = auxillary_post_location.stake_type
                tab['D{}'.format(rownumber)] = auxillary_post_location.azimuth
                tab['E{}'.format(rownumber)] = auxillary_post_location.distance

            i += 1


    def format_notes_tab(self, sheet, tab):
        tab['A1'] = 'Study Area'
        tab['B1'] = '=General!B1'

        tab['A2'] = 'Plot Number'
        tab['B2'] = '=General!B2'

        tab['A3'] = 'Deer Impact'
        tab['B3'] = '=General!B3'

        tab['A4'] = 'Date'
        tab['B4'] = '=General!B4'
        tab['B4'].number_format = 'M/D/YYYY'

        tab['A6'] = 'Deer Impact Logic'
        tab['A7'] = "(Please enter all notes into 'D' cells, no length constraint)"
        
        tab['A8'] = 'Seedlings'
        tab['B8'] = sheet.tabs[datasheet.TAB_NAME_NOTES].seedlings
        tab['C8'] = 'Notes:'
        tab['D8'] = sheet.tabs[datasheet.TAB_NAME_NOTES].seedlings_notes

        tab['A11'] = 'Browsing'
        tab['B11'] = sheet.tabs[datasheet.TAB_NAME_NOTES].browsing
        tab['C11'] = 'Notes:'
        tab['D11'] = sheet.tabs[datasheet.TAB_NAME_NOTES].browsing_notes

        tab['A14'] = 'Indicators'
        tab['B14'] = sheet.tabs[datasheet.TAB_NAME_NOTES].indicators
        tab['C14'] = 'Notes:'
        tab['D14'] = sheet.tabs[datasheet.TAB_NAME_NOTES].indicators_notes

        tab['A18'] = 'Notes:'
        tab['A19'] = sheet.tabs[datasheet.TAB_NAME_NOTES].general_notes


    def format_witness_trees_tab(self, sheet, tab):
        tab['A1'] = 'Witness Tree Table'

        tab['A2'] = 'Tree No'
        tab['B2'] = 'Subplot'
        tab['C2'] = 'Spp_K'
        tab['D2'] = 'Spp_G'
        tab['E2'] = 'dbh'
        tab['F2'] = 'L or D'
        tab['G2'] = 'Azimuth'
        tab['H2'] = 'Distance'
        tab['I2'] = 'UID'

        default_tree_number = 1
        i = 0

        for rownumber in range(3, 14):
            tab['A{}'.format(rownumber)] = default_tree_number

            if rownumber < 5 or (rownumber > 5 and rownumber < 7) or (rownumber > 7 and rownumber < 9) or (rownumber > 9 and rownumber < 11) or (rownumber > 11 and rownumber < 13):
                default_tree_number += 1
            else:
                default_tree_number = 1

            if i < len(sheet.tabs[datasheet.TAB_NAME_WITNESS_TREES].witness_trees):
                tree = sheet.tabs[datasheet.TAB_NAME_WITNESS_TREES].witness_trees[i]

                tab['B{}'.format(rownumber)] = tree.micro_plot_id
                tab['C{}'.format(rownumber)] = tree.species_known
                tab['D{}'.format(rownumber)] = tree.species_guess
                tab['E{}'.format(rownumber)] = tree.dbh
                tab['F{}'.format(rownumber)] = tree.live_or_dead
                tab['G{}'.format(rownumber)] = tree.azimuth
                tab['H{}'.format(rownumber)] = tree.distance
                tab['I{}'.format(rownumber)] = '=VALUE(CONCATENATE(General!$B$1,IF(LEN(General!$B$2)<2, CONCATENATE(0,General!$B$2),General!$B$2),IF(LEN(B3)<2, CONCATENATE(0,B3),B3)))'

            i += 1


    def format_cover_table_tab(self, sheet, tab):
        tab['A1'] = 'Cover Table'

        tab['A2'] = 'Micro'
        tab['B2'] = 'Quarter'
        tab['C2'] = '300/1000'
        tab['D2'] = 'Spp_K'
        tab['E2'] = 'Spp_G'
        tab['F2'] = 'PctCov'
        tab['G2'] = 'AvgHt'
        tab['H2'] = 'Count'
        tab['I2'] = 'Flower'
        tab['J2'] = 'Nstem'
        tab['K2'] = 'UID'

        i = 3
        for cover_species in sheet.tabs[datasheet.TAB_NAME_COVER_TABLE].cover_species:
            tab['A{}'.format(i)] = cover_species.micro_plot_id
            tab['B{}'.format(i)] = cover_species.quarter
            tab['C{}'.format(i)] = cover_species.scale
            tab['D{}'.format(i)] = cover_species.species_known
            tab['E{}'.format(i)] = cover_species.species_guess
            tab['F{}'.format(i)] = cover_species.percent_cover
            tab['G{}'.format(i)] = cover_species.average_height
            tab['H{}'.format(i)] = cover_species.count
            tab['I{}'.format(i)] = cover_species.flower
            tab['J{}'.format(i)] = cover_species.number_of_stems
            tab['K{}'.format(i)] = '=VALUE(CONCATENATE(General!$B$1,IF(LEN(General!$B$2)<2, CONCATENATE(0,General!$B$2),General!$B$2),IF(LEN(A{0})<2, CONCATENATE(0,A{0}),A{0})))'.format(i)

            i += 1


    def format_sapling_tab(self, sheet, tab):
        tab['A1'] = 'Sapling Table'

        tab['A2'] = 'Micro'
        tab['B2'] = 'Sapling No'
        tab['C2'] = 'Quarter'
        tab['D2'] = '300/1000'
        tab['E2'] = 'Spp_K'
        tab['F2'] = 'Spp_G'
        tab['G2'] = 'dbh'
        tab['H2'] = 'UID'

        i = 3
        for sapling_species in sheet.tabs[datasheet.TAB_NAME_SAPLING].sapling_species:
            tab['A{}'.format(i)] = sapling_species.micro_plot_id
            tab['B{}'.format(i)] = sapling_species.sapling_number
            tab['C{}'.format(i)] = sapling_species.quarter
            tab['D{}'.format(i)] = sapling_species.scale
            tab['E{}'.format(i)] = sapling_species.species_known
            tab['F{}'.format(i)] = sapling_species.species_guess
            tab['G{}'.format(i)] = sapling_species.diameter_breast_height
            tab['H{}'.format(i)] = '=VALUE(CONCATENATE(General!$B$1,IF(LEN(General!$B$2)<2, CONCATENATE(0,General!$B$2),General!$B$2),IF(LEN(A{0})<2, CONCATENATE(0,A{0}),A{0})))'.format(i)

            i += 1


    def format_seedling_tab(self, sheet, tab):
        tab['A1'] = 'Seedling Table'

        tab['A2'] = 'Micro'
        tab['B2'] = 'Quarter'
        tab['C2'] = '300/1000'
        tab['D2'] = 'Spp_K'
        tab['E2'] = 'Spp_G'
        tab['F2'] = 'Sprout'
        tab['G2'] = '0-6"'
        tab['H2'] = '6-12"'
        tab['I2'] = "1-3' Total"
        tab['J2'] = "1-3' Browsed"
        tab['K2'] = "3-5' Total"
        tab['L2'] = "3-5' Browsed"
        tab['M2'] = ">5' Total"
        tab['N2'] = ">5' Browsed"
        tab['O2'] = 'UID'

        i = 3
        for seedling_species in sheet.tabs[datasheet.TAB_NAME_SEEDLING].seedling_species:
            tab['A{}'.format(i)] = seedling_species.micro_plot_id
            tab['B{}'.format(i)] = seedling_species.quarter
            tab['C{}'.format(i)] = seedling_species.scale
            tab['D{}'.format(i)] = seedling_species.species_known
            tab['E{}'.format(i)] = seedling_species.species_guess
            tab['F{}'.format(i)] = seedling_species.sprout
            tab['G{}'.format(i)] = seedling_species.zero_six_inches
            tab['H{}'.format(i)] = seedling_species.six_twelve_inches
            tab['I{}'.format(i)] = seedling_species.one_three_feet_total
            tab['J{}'.format(i)] = seedling_species.one_three_feet_browsed
            tab['K{}'.format(i)] = seedling_species.three_five_feet_total
            tab['L{}'.format(i)] = seedling_species.three_five_feet_browsed
            tab['M{}'.format(i)] = seedling_species.greater_five_feet_total
            tab['N{}'.format(i)] = seedling_species.greater_five_feet_browsed
            tab['O{}'.format(i)] = '=VALUE(CONCATENATE(General!$B$1,IF(LEN(General!$B$2)<2, CONCATENATE(0,General!$B$2),General!$B$2),IF(LEN(A{0})<2, CONCATENATE(0,A{0}),A{0})))'.format(i)

            i += 1


    def format_tree_table_tab(self, sheet, tab):
        tab['A1'] = 'Tree Table'

        tab['A2'] = 'Subplot'
        tab['B2'] = 'Tree No'
        tab['C2'] = 'Spp_K'
        tab['D2'] = 'Spp_G'
        tab['E2'] = 'dbh'
        tab['F2'] = 'L or D'
        tab['G2'] = 'Comments'
        tab['H2'] = 'UID'

        i = 3
        for tree_species in sheet.tabs[datasheet.TAB_NAME_TREE_TABLE].tree_species:
            tab['A{}'.format(i)] = tree_species.micro_plot_id
            tab['B{}'.format(i)] = tree_species.tree_number
            tab['C{}'.format(i)] = tree_species.species_known
            tab['D{}'.format(i)] = tree_species.species_guess
            tab['E{}'.format(i)] = tree_species.diameter_breast_height
            tab['F{}'.format(i)] = tree_species.live_or_dead
            tab['G{}'.format(i)] = tree_species.comments
            tab['H{}'.format(i)] = '=VALUE(CONCATENATE(General!$B$1,IF(LEN(General!$B$2)<2, CONCATENATE(0,General!$B$2),General!$B$2),IF(LEN(A{0})<2, CONCATENATE(0,A{0}),A{0})))'.format(i)

            i += 1
