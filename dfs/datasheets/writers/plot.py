from openpyxl import Workbook
import dfs.datasheets.datasheet as datasheet
import re


class PlotDatasheetWriter:
    def write(self, sheet, output_directory):
        """
        Write a datasheet to an xlsx file.

        Keyword arguments:
        sheet -- a datasheet object.
        output_directory -- the directory to write the new xlsx file to.
        """

        workbook = Workbook()

        # remove the default worksheet
        workbook.remove(workbook['Sheet'])

        general_tab = workbook.create_sheet(title=datasheet.TAB_NAME_GENERAL)
        self.format_general_tab(sheet, general_tab)

        witness_tree_tab = workbook.create_sheet(title=datasheet.TAB_NAME_WITNESS_TREES)
        self.format_witness_trees_tab(sheet, witness_tree_tab)

        notes_tab = workbook.create_sheet(title=datasheet.TAB_NAME_NOTES)
        self.format_notes_tab(sheet, notes_tab)

        tree_table_tab = workbook.create_sheet(title=datasheet.TAB_NAME_TREE_TABLE)
        self.format_tree_table_tab(sheet, tree_table_tab)

        cover_table_tab = workbook.create_sheet(title=datasheet.TAB_NAME_COVER_TABLE)
        self.format_cover_table_tab(sheet, cover_table_tab)

        sapling_tab = workbook.create_sheet(title=datasheet.TAB_NAME_SAPLING)
        self.format_sapling_tab(sheet, sapling_tab)

        seedling_tab = workbook.create_sheet(title=datasheet.TAB_NAME_SEEDLING)
        self.format_seedling_tab(sheet, seedling_tab)

        workbook.save(f'{output_directory}/{sheet.output_filename}')


    def format_general_tab(self, sheet, tab):
        tab['A1'] = 'Study Area'
        tab['B1'] = int(sheet.tabs[datasheet.TAB_NAME_GENERAL].study_area)

        tab['A2'] = 'Plot Number'
        tab['B2'] = int(sheet.tabs[datasheet.TAB_NAME_GENERAL].plot_number)

        tab['A3'] = 'Deer Impact'
        tab['B3'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].deer_impact

        tab['A4'] = 'Date'
        tab['B4'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].collection_date
        tab['B4'].number_format = 'M/D/YYYY'

        tab['A6'] = 'Coordinate Converter'

        tab['A7'] = 'INPUT LAT HERE'
        tab['B7'] = 'INPUT LONG HERE'
        tab['C7'] = 'Data Type'
        tab['D7'] = 'Yes/No'
        tab['E7'] = 'Yes/No'
        tab['F7'] = '0-359'
        tab['G7'] = 'Feet'
        tab['H7'] = 'Meters'
        tab['I7'] = 'DO NOT INPUT VALUES HERE'

        tab['A8'] = 'ex. 4045.1291'
        tab['B8'] = 'ex. 7743.2763'
        tab['C8'] = 'Sub/ Micro'
        tab['D8'] = 'Collected'
        tab['E8'] = 'Fenced'
        tab['F8'] = 'Azimuth'
        tab['G8'] = 'Distance'
        tab['H8'] = 'Altitude'
        tab['I8'] = 'Latitude'
        tab['J8'] = 'Longitude'

        i = 0
        for rownumber in range(9, 14):
            if i >= len(sheet.tabs[datasheet.TAB_NAME_GENERAL].subplots):
                continue

            subplot = sheet.tabs[datasheet.TAB_NAME_GENERAL].subplots[i]

            tab[f'A{rownumber}'] = subplot.latitude
            tab[f'B{rownumber}'] = subplot.longitude
            tab[f'C{rownumber}'] = int(subplot.micro_plot_id)

            if None == subplot.collected:
                if subplot.micro_plot_id in sheet.tabs[datasheet.TAB_NAME_COVER_TABLE].get_recorded_subplots():
                    subplot.collected = 'Yes'
                else:
                    subplot.collected = 'No'

            tab[f'D{rownumber}'] = subplot.collected
            tab[f'E{rownumber}'] = subplot.fenced
            tab[f'F{rownumber}'] = subplot.azimuth
            tab[f'G{rownumber}'] = subplot.distance
            tab[f'H{rownumber}'] = subplot.altitude

            if subplot.converted_latitude:
                tab[f'I{rownumber}'] = subplot.converted_latitude
            else:
                tab[f'I{rownumber}'] = '=IF(ISBLANK(A{0}),"",LEFT((LEFT(A{0},2)+(RIGHT(A{0},LEN(A{0})-2)/60)),10))'.format(rownumber)

            if subplot.converted_longitude:
                tab[f'J{rownumber}'] = subplot.converted_longitude
            else:
                tab[f'J{rownumber}'] = '=IF(ISBLANK(B{0}),"",LEFT(-1*(LEFT(B{0},2)+(RIGHT(B{0},LEN(B{0})-2)/60)),10))'.format(rownumber)

            i += 1

        tab['A15'] = 'Data Type'
        tab['B15'] = 'Yes/No'
        tab['C15'] = 'Yes/No'
        tab['D15'] = '0-1'
        tab['E15'] = 'Type'
        tab['F15'] = 'Yes/No'
        tab['G15'] = 'Yes/No'
        tab['H15'] = "(notes contained in 'H' cells only)"

        tab['A16'] = 'Sub/Micro'
        tab['B16'] = 'Re-Monumented'
        tab['C16'] = 'Forested'
        tab['D16'] = 'Disturbance'
        tab['E16'] = 'Type'
        tab['F16'] = 'Lime'
        tab['G16'] = 'Herbicide'
        tab['H16'] = 'Notes'

        i = 0
        for rownumber in range(17, 22):
            if i >= len(sheet.tabs[datasheet.TAB_NAME_GENERAL].subplots):
                continue

            subplot = sheet.tabs[datasheet.TAB_NAME_GENERAL].subplots[i]

            tab[f'A{rownumber}'] = subplot.micro_plot_id
            tab[f'B{rownumber}'] = subplot.re_monumented
            tab[f'C{rownumber}'] = subplot.forested
            tab[f'D{rownumber}'] = subplot.disturbance
            tab[f'E{rownumber}'] = subplot.disturbance_type
            tab[f'F{rownumber}'] = subplot.lime
            tab[f'G{rownumber}'] = subplot.herbicide
            tab[f'H{rownumber}'] = subplot.notes

            i += 1

        tab['A23'] = 'Fenced Subplot Condition'
        tab['C23'] = '0-3'
        tab['D23'] = '(notes contained in D25 only)'

        tab['A24'] = 'Active Exclosure'
        tab['B24'] = 'Repair(s)'
        tab['C24'] = 'Level'
        tab['D24'] = 'Notes'

        tab['A25'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].fenced_subplot_condition.active_exclosure
        tab['B25'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].fenced_subplot_condition.repairs
        tab['C25'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].fenced_subplot_condition.level
        tab['D25'] = sheet.tabs[datasheet.TAB_NAME_GENERAL].fenced_subplot_condition.notes

        tab['A27'] = 'Auxillary Post Locations'
        tab['A28'] = '1-5'
        tab['B28'] = '1 or 2'
        tab['C28'] = 'Wooden or Rebar'
        tab['D28'] = '0-359'
        tab['E28'] = 'Feet'

        tab['A29'] = 'Sub/Micro'
        tab['B29'] = 'Post'
        tab['C29'] = 'Stake Type'
        tab['D29'] = 'Azimuth'
        tab['E29'] = 'Distance'

        i = 0
        for rownumber in range(30, 35):
            if i < len(sheet.tabs[datasheet.TAB_NAME_GENERAL].auxillary_post_locations):
                auxillary_post_location = sheet.tabs[datasheet.TAB_NAME_GENERAL].auxillary_post_locations[i]

                tab[f'A{rownumber}'] = auxillary_post_location.micro_plot_id
                tab[f'B{rownumber}'] = auxillary_post_location.post
                tab[f'C{rownumber}'] = auxillary_post_location.stake_type
                tab[f'D{rownumber}'] = auxillary_post_location.azimuth
                tab[f'E{rownumber}'] = auxillary_post_location.distance
            else:
                tab[f'A{rownumber}'] = ''
                tab[f'B{rownumber}'] = ''
                tab[f'C{rownumber}'] = ''
                tab[f'D{rownumber}'] = ''
                tab[f'E{rownumber}'] = ''

            i += 1

        tab['A36'] = 'Non-Forest Azimuths'
        tab['C36'] = '(if a subplot is not 100% forested, record 3 azimuths along the forested barrier from subplot center)'

        tab['A37'] = 'Subplot'
        tab['B37'] = 'Azimuth 1'
        tab['C37'] = 'Azimuth 2'
        tab['D37'] = 'Azimuth 3'

        i = 0
        for rownumber in range(38, 42):
            if i < len(sheet.tabs[datasheet.TAB_NAME_GENERAL].non_forested_azimuths):
                non_forested_azimuth = sheet.tabs[datasheet.TAB_NAME_GENERAL].non_forested_azimuths[i]

                tab[f'A{rownumber}'] = non_forested_azimuth.micro_plot_id
                tab[f'B{rownumber}'] = non_forested_azimuth.azimuth_1
                tab[f'C{rownumber}'] = non_forested_azimuth.azimuth_2
                tab[f'D{rownumber}'] = non_forested_azimuth.azimuth_3
            else:
                tab[f'A{rownumber}'] = ''
                tab[f'B{rownumber}'] = ''
                tab[f'C{rownumber}'] = ''
                tab[f'D{rownumber}'] = ''



    def format_notes_tab(self, sheet, tab):
        tab['A1'] = 'Study Area'
        tab['C1'] = '=General!B1'

        tab['A2'] = 'Plot Number'
        tab['C2'] = '=General!B2'

        tab['A3'] = 'Deer Impact'
        tab['C3'] = '=General!B3'

        tab['A4'] = 'Date'
        tab['C4'] = '=General!B4'
        tab['C4'].number_format = 'M/D/YYYY'

        tab['A6'] = 'Deer Impact Logic'
        tab['A7'] = "(Please enter all notes into 'D' cells, no length constraint)"
        
        tab['A8'] = 'Seedlings'
        tab['B8'] = sheet.tabs[datasheet.TAB_NAME_NOTES].seedlings
        tab['C8'] = 'Notes:'
        tab['D8'] = sheet.tabs[datasheet.TAB_NAME_NOTES].seedlings_notes

        tab['A11'] = 'Browsing'
        tab['B11'] = sheet.tabs[datasheet.TAB_NAME_NOTES].browsing
        tab['C11'] = 'Notes:'
        tab['D11'] = sheet.tabs[datasheet.TAB_NAME_NOTES].browsing_notes

        tab['A14'] = 'Indicators'
        tab['B14'] = sheet.tabs[datasheet.TAB_NAME_NOTES].indicators
        tab['C14'] = 'Notes:'
        tab['D14'] = sheet.tabs[datasheet.TAB_NAME_NOTES].indicators_notes

        tab['A18'] = 'Notes:'
        tab['A19'] = sheet.tabs[datasheet.TAB_NAME_NOTES].general_notes


    def format_witness_trees_tab(self, sheet, tab):
        tab['A1'] = 'Witness Tree Table'

        tab['A2'] = 'Tree No'
        tab['B2'] = 'Subplot'
        tab['C2'] = 'Spp_K'
        tab['D2'] = 'Spp_G'
        tab['E2'] = 'dbh'
        tab['F2'] = 'L or D'
        tab['G2'] = 'Azimuth'
        tab['H2'] = 'Distance'

        default_tree_number = 1
        i = 0

        for rownumber in range(3, 14):
            tab[f'A{rownumber}'] = default_tree_number

            if rownumber < 5 or (rownumber > 5 and rownumber < 7) or (rownumber > 7 and rownumber < 9) or (rownumber > 9 and rownumber < 11) or (rownumber > 11 and rownumber < 13):
                default_tree_number += 1
            else:
                default_tree_number = 1

            if i < len(sheet.tabs[datasheet.TAB_NAME_WITNESS_TREES].witness_trees):
                tree = sheet.tabs[datasheet.TAB_NAME_WITNESS_TREES].witness_trees[i]

                tab[f'B{rownumber}'] = tree.micro_plot_id
                tab[f'C{rownumber}'] = tree.species_known
                tab[f'D{rownumber}'] = tree.species_guess
                tab[f'E{rownumber}'] = tree.dbh
                tab[f'F{rownumber}'] = tree.live_or_dead
                tab[f'G{rownumber}'] = tree.azimuth
                tab[f'H{rownumber}'] = tree.distance
            else:
                tab[f'B{rownumber}'] = ''
                tab[f'C{rownumber}'] = ''
                tab[f'D{rownumber}'] = ''
                tab[f'E{rownumber}'] = ''
                tab[f'F{rownumber}'] = ''
                tab[f'G{rownumber}'] = ''
                tab[f'H{rownumber}'] = ''

            i += 1


    def format_cover_table_tab(self, sheet, tab):
        tab['A1'] = 'Cover Table'

        tab['A2'] = 'Micro'
        tab['B2'] = 'Quarter'
        tab['C2'] = '300/1000'
        tab['D2'] = 'Spp_K'
        tab['E2'] = 'Spp_G'
        tab['F2'] = 'PctCov'
        tab['G2'] = 'AvgHt'
        tab['H2'] = 'Count'
        tab['I2'] = 'Flower'
        tab['J2'] = 'Nstem'

        i = 3

        if 0 == len(sheet.tabs[datasheet.TAB_NAME_COVER_TABLE].cover_species):
            tab[f'A{i}'] = ''
            tab[f'B{i}'] = ''
            tab[f'C{i}'] = ''
            tab[f'D{i}'] = ''
            tab[f'E{i}'] = ''
            tab[f'F{i}'] = ''
            tab[f'G{i}'] = ''
            tab[f'H{i}'] = ''
            tab[f'I{i}'] = ''
            tab[f'J{i}'] = ''
        else:
            for cover_species in sheet.tabs[datasheet.TAB_NAME_COVER_TABLE].cover_species:
                tab[f'A{i}'] = cover_species.micro_plot_id
                tab[f'B{i}'] = cover_species.quarter
                tab[f'C{i}'] = cover_species.scale
                tab[f'D{i}'] = cover_species.species_known
                tab[f'E{i}'] = cover_species.species_guess
                tab[f'F{i}'] = cover_species.percent_cover
                tab[f'G{i}'] = cover_species.average_height
                tab[f'H{i}'] = cover_species.count
                tab[f'I{i}'] = cover_species.flower
                tab[f'J{i}'] = cover_species.number_of_stems

                i += 1


    def format_sapling_tab(self, sheet, tab):
        tab['A1'] = 'Sapling Table'

        tab['A2'] = 'Micro'
        tab['B2'] = 'Sapling No'
        tab['C2'] = 'Quarter'
        tab['D2'] = '300/1000'
        tab['E2'] = 'Spp_K'
        tab['F2'] = 'Spp_G'
        tab['G2'] = 'dbh'

        i = 3

        if 0 == len(sheet.tabs[datasheet.TAB_NAME_SAPLING].sapling_species):
            tab[f'A{i}'] = ''
            tab[f'B{i}'] = ''
            tab[f'C{i}'] = ''
            tab[f'D{i}'] = ''
            tab[f'E{i}'] = ''
            tab[f'F{i}'] = ''
            tab[f'G{i}'] = ''
        else: 
            for sapling_species in sheet.tabs[datasheet.TAB_NAME_SAPLING].sapling_species:
                tab[f'A{i}'] = sapling_species.micro_plot_id
                tab[f'B{i}'] = sapling_species.sapling_number
                tab[f'C{i}'] = sapling_species.quarter
                tab[f'D{i}'] = sapling_species.scale
                tab[f'E{i}'] = sapling_species.species_known
                tab[f'F{i}'] = sapling_species.species_guess
                tab[f'G{i}'] = sapling_species.diameter_breast_height

                i += 1


    def format_seedling_tab(self, sheet, tab):
        tab['A1'] = 'Seedling Table'

        tab['A2'] = 'Micro'
        tab['B2'] = 'Quarter'
        tab['C2'] = '300/1000'
        tab['D2'] = 'Spp_K'
        tab['E2'] = 'Spp_G'
        tab['F2'] = 'Sprout'
        tab['G2'] = '0-6"'
        tab['H2'] = '6-12"'
        tab['I2'] = "1-3' Total"
        tab['J2'] = "1-3' Browsed"
        tab['K2'] = "3-5' Total"
        tab['L2'] = "3-5' Browsed"
        tab['M2'] = ">5' Total"
        tab['N2'] = ">5' Browsed"

        i = 3
        if 0 == len(sheet.tabs[datasheet.TAB_NAME_SEEDLING].seedling_species):
            tab[f'A{i}'] = ''
            tab[f'B{i}'] = ''
            tab[f'C{i}'] = ''
            tab[f'D{i}'] = ''
            tab[f'E{i}'] = ''
            tab[f'F{i}'] = ''
            tab[f'G{i}'] = ''
            tab[f'H{i}'] = ''
            tab[f'I{i}'] = ''
            tab[f'J{i}'] = ''
            tab[f'K{i}'] = ''
            tab[f'L{i}'] = ''
            tab[f'M{i}'] = ''
            tab[f'N{i}'] = ''
        else:
            for seedling_species in sheet.tabs[datasheet.TAB_NAME_SEEDLING].seedling_species:
                tab[f'A{i}'] = seedling_species.micro_plot_id
                tab[f'B{i}'] = seedling_species.quarter
                tab[f'C{i}'] = seedling_species.scale
                tab[f'D{i}'] = seedling_species.species_known
                tab[f'E{i}'] = seedling_species.species_guess
                tab[f'F{i}'] = seedling_species.sprout
                tab[f'G{i}'] = seedling_species.zero_six_inches
                tab[f'H{i}'] = seedling_species.six_twelve_inches
                tab[f'I{i}'] = seedling_species.one_three_feet_total
                tab[f'J{i}'] = seedling_species.one_three_feet_browsed
                tab[f'K{i}'] = seedling_species.three_five_feet_total
                tab[f'L{i}'] = seedling_species.three_five_feet_browsed
                tab[f'M{i}'] = seedling_species.greater_five_feet_total
                tab[f'N{i}'] = seedling_species.greater_five_feet_browsed

                i += 1


    def format_tree_table_tab(self, sheet, tab):
        tab['A1'] = 'Tree Table'

        tab['A2'] = 'Subplot'
        tab['B2'] = 'Tree No'
        tab['C2'] = 'Spp_K'
        tab['D2'] = 'Spp_G'
        tab['E2'] = 'dbh'
        tab['F2'] = 'L or D'
        tab['G2'] = 'Comments'

        i = 3
        if 0 == len(sheet.tabs[datasheet.TAB_NAME_TREE_TABLE].tree_species):
            tab[f'A{i}'] = ''
            tab[f'B{i}'] = ''
            tab[f'C{i}'] = ''
            tab[f'D{i}'] = ''
            tab[f'E{i}'] = ''
            tab[f'F{i}'] = ''
            tab[f'G{i}'] = ''
        else:
            for tree_species in sheet.tabs[datasheet.TAB_NAME_TREE_TABLE].tree_species:
                tab[f'A{i}'] = tree_species.micro_plot_id
                tab[f'B{i}'] = tree_species.tree_number
                tab[f'C{i}'] = tree_species.species_known
                tab[f'D{i}'] = tree_species.species_guess
                tab[f'E{i}'] = tree_species.diameter_breast_height
                tab[f'F{i}'] = tree_species.live_or_dead
                tab[f'G{i}'] = tree_species.comments

                i += 1
